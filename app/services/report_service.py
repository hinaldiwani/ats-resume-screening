"""
app/services/report_service.py

Module 11 — Reports. Generates a PDF and an Excel export of a job's
ranked screening results, reusing screening_service's ranking and
skill-match logic directly rather than recomputing anything separately —
so the report always reflects exactly what the API/UI show, with no
second, potentially-drifting data path.
"""

import io
import logging
from datetime import datetime
from typing import List

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.models.models import JobDescription, ATSScore
from app.services import scoring_service

logger = logging.getLogger(__name__)


def generate_pdf_report(job: JobDescription, ats_scores: List[ATSScore]) -> bytes:
    """Builds a ranked-candidates PDF for one job. Returns raw PDF bytes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"ATS Screening Report: {job.title}", styles["Title"]))
    elements.append(Paragraph(
        f"Generated {datetime.utcnow().strftime('%B %d, %Y')} &middot; {len(ats_scores)} candidate(s) screened",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 0.2 * inch))

    if job.description:
        elements.append(Paragraph("Job Description", styles["Heading2"]))
        truncated = job.description[:500] + ("..." if len(job.description) > 500 else "")
        elements.append(Paragraph(truncated, styles["Normal"]))
        elements.append(Spacer(1, 0.25 * inch))

    if not ats_scores:
        elements.append(Paragraph("No candidates have been screened for this job yet.", styles["Normal"]))
    else:
        table_data = [["Rank", "Candidate", "Email", "Final Score", "Skill Match", "Experience Match", "Status"]]
        for i, score in enumerate(ats_scores, start=1):
            table_data.append([
                str(i),
                score.resume.candidate.name,
                score.resume.candidate.email,
                f"{score.final_score:.1f}",
                f"{score.skill_match_score:.1f}%",
                f"{score.experience_match_score:.1f}%",
                score.status.capitalize(),
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A1120")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F8")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD2DE")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def generate_excel_report(job: JobDescription, ats_scores: List[ATSScore]) -> bytes:
    """Builds a ranked-candidates Excel workbook for one job. Returns raw XLSX bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    ws["A1"] = f"ATS Screening Report: {job.title}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Generated {datetime.utcnow().strftime('%B %d, %Y')} \u2014 {len(ats_scores)} candidate(s) screened"
    ws["A2"].font = Font(italic=True, color="666666")

    headers = [
        "Rank", "Candidate", "Email", "Final Score", "Skill Match %",
        "Experience Match %", "Matched Skills", "Missing Skills", "Status",
    ]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0E7C66", end_color="0E7C66", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for i, score in enumerate(ats_scores, start=1):
        skill_result = scoring_service.calculate_skill_match(score.resume.skills, score.job_description.skills)
        row = header_row + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=score.resume.candidate.name)
        ws.cell(row=row, column=3, value=score.resume.candidate.email)
        ws.cell(row=row, column=4, value=score.final_score)
        ws.cell(row=row, column=5, value=score.skill_match_score)
        ws.cell(row=row, column=6, value=score.experience_match_score)
        ws.cell(row=row, column=7, value=", ".join(skill_result["matched_skills"]))
        ws.cell(row=row, column=8, value=", ".join(skill_result["missing_skills"]))
        ws.cell(row=row, column=9, value=score.status.capitalize())

    column_widths = [6, 20, 28, 12, 14, 18, 34, 34, 12]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
