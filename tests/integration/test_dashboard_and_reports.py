"""
tests/integration/test_dashboard_and_reports.py

Integration tests covering Module 9 (Dashboard stats & aggregations) and
Module 11 (PDF and Excel report generation).
"""

import io

from openpyxl import load_workbook
from reportlab.pdfgen import canvas



def make_pdf_bytes(lines):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer)
    y = 750
    for line in lines:
        c.drawString(72, y, line)
        y -= 18
    c.save()
    buffer.seek(0)
    return buffer.read()


RESUME_LINES = [
    "Sara Connor", "sara@example.com",
    "", "SKILLS", "Python, FastAPI, Docker, Machine Learning",
    "", "EXPERIENCE", "5 years of experience in AI engineering.",
    "", "EDUCATION", "M.S. in Computer Science, Stanford",
]


def test_dashboard_stats_endpoint(client, auth_headers):
    # Create a job first
    r_job = client.post("/api/v1/jobs/", json={
        "title": "AI Engineer",
        "description": "Building AI services with python fastapi docker",
        "min_experience_years": 3.0,
        "skills": ["Python", "FastAPI", "Docker"],
    }, headers=auth_headers)
    assert r_job.status_code == 201

    # Fetch stats
    r_stats = client.get("/api/v1/dashboard/stats", headers=auth_headers)
    assert r_stats.status_code == 200
    data = r_stats.json()["data"]
    assert len(data) == 4
    labels = {item["label"] for item in data}
    assert labels == {"Open positions", "Candidates screened", "Avg. match score", "Awaiting review"}


def test_dashboard_top_candidates_and_recent_uploads(client, auth_headers):
    # Upload resume
    pdf_bytes = make_pdf_bytes(RESUME_LINES)
    r_upload = client.post(
        "/api/v1/resumes/upload",
        files={"file": ("sara_resume.pdf", pdf_bytes, "application/pdf")},
        headers=auth_headers,
    )
    assert r_upload.status_code == 201

    # Create job
    r_job = client.post("/api/v1/jobs/", json={
        "title": "ML Engineer",
        "description": "Python, FastAPI, Docker machine learning",
        "skills": ["Python", "FastAPI", "Docker"],
    }, headers=auth_headers)
    job_id = r_job.json()["data"]["id"]
    resume_id = r_upload.json()["data"]["id"]

    # Run screening
    client.post("/api/v1/screening/run", json={"resume_id": resume_id, "job_id": job_id}, headers=auth_headers)

    # Check top candidates
    r_top = client.get("/api/v1/dashboard/top-candidates", headers=auth_headers)
    assert r_top.status_code == 200
    top_candidates = r_top.json()["data"]
    assert len(top_candidates) >= 1
    assert top_candidates[0]["name"] == "Sara Connor"

    # Check recent uploads
    r_uploads = client.get("/api/v1/dashboard/recent-uploads", headers=auth_headers)
    assert r_uploads.status_code == 200
    uploads = r_uploads.json()["data"]
    assert len(uploads) >= 1
    assert uploads[0]["candidate_name"] == "Sara Connor"


def test_generate_pdf_report_endpoint(client, auth_headers):
    # Create job
    r_job = client.post("/api/v1/jobs/", json={
        "title": "Lead Backend Dev",
        "description": "Python FastAPI PostgreSQL Docker",
        "skills": ["Python", "FastAPI"],
    }, headers=auth_headers)
    job_id = r_job.json()["data"]["id"]

    # Upload resume & screen
    pdf_bytes = make_pdf_bytes(RESUME_LINES)
    r_up = client.post(
        "/api/v1/resumes/upload",
        files={"file": ("resume.pdf", pdf_bytes, "application/pdf")},
        headers=auth_headers,
    )
    resume_id = r_up.json()["data"]["id"]
    client.post("/api/v1/screening/run", json={"resume_id": resume_id, "job_id": job_id}, headers=auth_headers)

    # Download PDF
    r_pdf = client.get(f"/api/v1/reports/jobs/{job_id}/pdf", headers=auth_headers)
    assert r_pdf.status_code == 200
    assert r_pdf.headers["content-type"] == "application/pdf"
    assert r_pdf.content.startswith(b"%PDF")


def test_generate_excel_report_endpoint(client, auth_headers):
    # Create job
    r_job = client.post("/api/v1/jobs/", json={
        "title": "Staff Engineer",
        "description": "Python FastAPI",
        "skills": ["Python"],
    }, headers=auth_headers)
    job_id = r_job.json()["data"]["id"]

    # Download Excel report
    r_excel = client.get(f"/api/v1/reports/jobs/{job_id}/excel", headers=auth_headers)
    assert r_excel.status_code == 200
    assert "spreadsheetml" in r_excel.headers["content-type"]

    # Verify openpyxl can load the generated workbook
    wb = load_workbook(io.BytesIO(r_excel.content))
    assert "Screening Results" in wb.sheetnames
